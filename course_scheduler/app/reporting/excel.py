"""
Excel reporting utilities for schedule export.

This module provides functions for exporting schedules and analysis data
to Excel format with multiple sheets and formatting.
"""
import logging
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..data.models import Schedule

logger = logging.getLogger(__name__)


def export_schedules_to_excel(schedules: List[Schedule], output_path: str,
                             include_analysis: bool = True) -> bool:
    """
    Export schedules to Excel with multiple sheets.

    Args:
        schedules: List of Schedule objects to export
        output_path: Path where Excel file will be saved
        include_analysis: Whether to include analysis sheets

    Returns:
        True if successful, False otherwise
    """
    try:
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Schedule summary sheet
        ws_summary = wb.create_sheet("Schedule Summary")
        _create_summary_sheet(ws_summary, schedules)

        # Individual schedule sheets
        for i, schedule in enumerate(schedules, 1):
            sheet_name = f"Schedule_{i}"
            ws_schedule = wb.create_sheet(sheet_name)
            _create_schedule_sheet(ws_schedule, schedule, i)

        # Analysis sheet if requested
        if include_analysis and schedules:
            ws_analysis = wb.create_sheet("Analysis")
            _create_analysis_sheet(ws_analysis, schedules)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Successfully exported {len(schedules)} schedules to Excel: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Error exporting schedules to Excel: {e}")
        return False


def _create_summary_sheet(worksheet, schedules: List[Schedule]):
    """Create the summary sheet with overview of all schedules."""
    # Headers
    headers = ["Schedule #", "Total Credits", "Course Count", "Conflicts", "Total Hours"]
    worksheet.append(headers)

    # Format headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, schedule in enumerate(schedules, 1):
        total_hours = len(set((day, slot) for course in schedule.courses for day, slot in course.schedule))
        row_data = [i, schedule.total_credits, len(schedule.courses), schedule.conflict_count, total_hours]
        worksheet.append(row_data)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _create_schedule_sheet(worksheet, schedule: Schedule, schedule_num: int):
    """Create a detailed sheet for a single schedule."""
    # Title
    worksheet.append([f"Schedule {schedule_num} Details"])
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.font = Font(size=16, bold=True)

    # Summary info
    worksheet.append([])
    worksheet.append(["Total Credits:", schedule.total_credits])
    worksheet.append(["Number of Courses:", len(schedule.courses)])
    worksheet.append(["Conflicts:", schedule.conflict_count])
    worksheet.append([])

    # Course details headers
    headers = ["Course Code", "Course Name", "Credits", "Type", "Schedule", "Instructor", "Faculty", "Department"]
    worksheet.append(headers)

    # Format headers
    header_row = worksheet.max_row
    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Course data
    for course in schedule.courses:
        schedule_str = ", ".join(f"{day}{slot}" for day, slot in course.schedule)
        row_data = [
            course.code,
            course.name,
            course.ects,
            course.course_type,
            schedule_str,
            course.teacher,
            getattr(course, 'faculty', ''),
            getattr(course, 'department', '')
        ]
        worksheet.append(row_data)

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _create_analysis_sheet(worksheet, schedules: List[Schedule]):
    """Create analysis sheet with statistics."""
    # Title
    worksheet.append(["Schedule Analysis"])
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.font = Font(size=16, bold=True)

    worksheet.append([])

    # Basic statistics
    total_schedules = len(schedules)
    avg_credits = sum(s.total_credits for s in schedules) / total_schedules
    avg_courses = sum(len(s.courses) for s in schedules) / total_schedules
    conflict_free = sum(1 for s in schedules if s.conflict_count == 0)

    stats_data = [
        ["Total Schedules Generated:", total_schedules],
        ["Average Credits per Schedule:", f"{avg_credits:.1f}"],
        ["Average Courses per Schedule:", f"{avg_courses:.1f}"],
        ["Conflict-Free Schedules:", f"{conflict_free} ({conflict_free/total_schedules*100:.1f}%)"],
        []
    ]

    for row_data in stats_data:
        worksheet.append(row_data)

    # Course frequency analysis
    worksheet.append(["Most Frequently Selected Courses:"])
    course_frequency = {}
    for schedule in schedules:
        for course in schedule.courses:
            course_frequency[course.code] = course_frequency.get(course.code, 0) + 1

    # Sort by frequency
    sorted_courses = sorted(course_frequency.items(), key=lambda x: x[1], reverse=True)

    worksheet.append(["Course Code", "Frequency", "Percentage"])
    header_row = worksheet.max_row
    for col in range(1, 4):
        cell = worksheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for course_code, freq in sorted_courses[:20]:  # Top 20
        percentage = (freq / total_schedules) * 100
        worksheet.append([course_code, freq, f"{percentage:.1f}%"])


def export_schedule_analysis_to_excel(analysis_data: Dict[str, Any], output_path: str) -> bool:
    """
    Export schedule analysis data to Excel.

    Args:
        analysis_data: Dictionary containing analysis results
        output_path: Path where Excel file will be saved

    Returns:
        True if successful, False otherwise
    """
    try:
        # Create DataFrame from analysis data
        df = pd.DataFrame([analysis_data])

        # Save to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Analysis Summary', index=False)

        logger.info(f"Successfully exported analysis to Excel: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Error exporting analysis to Excel: {e}")
        return False
